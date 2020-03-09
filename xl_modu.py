from openpyxl import Workbook
from openpyxl.chart import AreaChart3D, BarChart3D, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import colors
from openpyxl.styles import Font
import datetime as dt


preds = ['today', 'tomorrow', '2_days', '3_days', '4_days', '5_days', '6_days', '1_week', '2_weeks', 'month']
today = str(dt.datetime.today().date())


def process_workbook(stock_names, predictions, prices, r2_scores, scores, score_avgs, r2_avgs):
    # Creating the xl file with the analyzed data
    wb = Workbook()
    ws = wb.create_sheet('Home-Page')
    ws.append(['Stock', 'R2_avgs', 'Scores_avgs'])
    for i, stock in enumerate(stock_names):
        ws1 = wb.create_sheet(stock)
        ws1.append(['Days', 'Values', 'R2_Scores', 'My_Scores'])
        ws1.append(['Price', prices[i], 1, 0])
        for j, raw in enumerate(preds):
            ws1.append([raw, predictions[j][i], r2_scores[j][i], scores[j][i]])

        # Creating and adding the charts of the analyzed data
        titles = Reference(ws1, min_col=1, min_row=2, max_row=12, max_col=1)
        data_val = Reference(ws1, min_col=2, min_row=1, max_col=2, max_row=12)
        data_r2_score = Reference(ws1, min_col=3, min_row=1, max_col=3, max_row=12)
        data_my_score = Reference(ws1, min_col=4, min_row=1, max_col=4, max_row=12)
        chart_1 = create_chart(AreaChart3D(), 'Predictions', 'Values', titles, data_val, stock, "y")
        chart_2 = create_chart(BarChart3D(), 'Predictions', 'R2_Scores', titles, data_r2_score, stock, "b")
        chart_3 = create_chart(BarChart3D(), 'Predictions', 'My_Scores', titles, data_my_score, stock, "r")
        ws1.add_chart(chart_1, "F2")
        ws1.add_chart(chart_2, "F25")
        ws1.add_chart(chart_3, "P2")

        # Add a table
        tab = Table(displayName="Table1", ref="A1:D12")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws1.add_table(tab)

        # Add hyperlinks with font between the current sheet to the home page sheet
        ft = Font(color=colors.RED)
        ws1.cell(row=len(preds) + 4, column=2).value = "Home-Page"
        ws1.cell(row=len(preds) + 4, column=2).font = ft
        ws1.cell(row=len(preds) + 4, column=2).hyperlink = '#\'Home-Page\'!A1'
        ws.append(['  '+stock_names[i], r2_avgs[i], score_avgs[i]])
        ws.cell(row=i+2, column=1).hyperlink = f'#{stock_names[i]}!A1'
        ws.cell(row=i + 2, column=1).font = ft

    std = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(std)
    set_home_page(ws, len(stock_names))
    wb.save(f"xl_Static_analyzes/Data_Anlaz_{today}.xlsx")
    # wb.save(f"xl_Static_analyzes/Test_{today}.xlsx")


# General function to create chart
def create_chart(ch_type, x_ti, y_ti, x_vals, y_vals, stock_name, color):
    chart = ch_type
    chart.x_axis.title = x_ti
    chart.y_axis.title = y_ti
    chart.title = stock_name + ' Predictions - ' + y_ti
    chart.add_data(data=y_vals, titles_from_data=True)
    chart.set_categories(x_vals)
    chart.height = 10.5  # default is 7.5
    chart.width = 18.5  # default is 15

    if color != "b":
        set_color(chart, color)
    else:
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = "00000"
    return chart


# Function to set the color of the chart
def set_color(chart, color):
    col = {'y': "ff9900",
           'r': "FF0000",
           'b': "00000"}
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = col['b']
    s.graphicalProperties.solidFill = col[color]


# Function the set the charts and table of the home page
def set_home_page(sheet, num_stocks):
    titles = Reference(sheet, min_col=1, min_row=2, max_row=num_stocks+1, max_col=1)
    r2_val = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=num_stocks+1)
    score_val = Reference(sheet, min_col=3, min_row=1, max_col=3, max_row=num_stocks+1)

    chart_1 = create_chart(BarChart3D(), 'Stock names', 'R2_Averages', titles, r2_val, "Model", "b")
    chart_2 = create_chart(BarChart3D(), 'Stock names', 'Scores_Averages', titles, score_val, "Model", "r")
    chart_1.width = 40  # default is 18.5
    chart_2.width = 40  # default is 18.5
    sheet.add_chart(chart_1, "F2")
    sheet.add_chart(chart_2, "F25")

    tab = Table(displayName="Table1", ref=f"A1:C{num_stocks+1}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True,
                           showColumnStripes=True)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
